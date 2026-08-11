from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ledger_reporter.domain.models import ReportBundle
from ledger_reporter.presentation.builders import build_tables
from ledger_reporter.presentation.models import TableSpec
from ledger_reporter.presentation.theme import STYLES


def _excel_value(value: object) -> object:
    return float(value) if isinstance(value, Decimal) else value


def _configure_sheet(sheet: Worksheet, table: TableSpec) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B3" if table.name == "经营汇总" else "A3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:2"


def _write_table(sheet: Worksheet, table: TableSpec) -> None:
    thin = Side(style="thin", color="B7C0BB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell_spec in table.cells:
        cell = sheet.cell(cell_spec.row, cell_spec.column, _excel_value(cell_spec.value))
        style = STYLES[cell_spec.style]
        cell.fill = PatternFill("solid", fgColor=style["fill"])
        cell.font = Font(
            name="Arial",
            size=10,
            bold=style["bold"],
            color=style.get("font", "000000"),
        )
        cell.alignment = Alignment(horizontal=style["align"], vertical="center")
        cell.border = border
        if cell_spec.number_format:
            cell.number_format = cell_spec.number_format

    for merge in table.merges:
        sheet.merge_cells(
            start_row=merge.start_row,
            start_column=merge.start_column,
            end_row=merge.end_row,
            end_column=merge.end_column,
        )
    for index, width in enumerate(table.column_widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for index, height in enumerate(table.row_heights, 1):
        sheet.row_dimensions[index].height = height


def export_excel(bundle: ReportBundle, output: Path) -> None:
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        workbook.remove(workbook.active)
        for table in build_tables(bundle):
            sheet = workbook.create_sheet(table.name)
            _configure_sheet(sheet, table)
            _write_table(sheet, table)
        workbook.save(output)
    finally:
        workbook.close()
