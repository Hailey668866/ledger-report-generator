from collections.abc import Callable, Iterable, Iterator
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from ledger_reporter.domain.models import FundRecord, OperationalRecord
from ledger_reporter.io.errors import WorkbookDataError
from ledger_reporter.io.source_settings import DEFAULT_SOURCE_SETTINGS, SourceSettings


def _open_workbook(path: Path, data_only: bool):
    source = Path(path)
    if source.suffix.lower() != ".xlsx":
        raise WorkbookDataError(f"文件「{source}」不是 .xlsx 工作簿。")
    if not source.is_file():
        raise WorkbookDataError(f"找不到工作簿文件「{source}」。")
    try:
        return load_workbook(source, read_only=True, data_only=data_only, keep_links=False)
    except (BadZipFile, InvalidFileException, KeyError):
        raise WorkbookDataError(f"文件「{source}」不是有效的 XLSX 工作簿。") from None
    except OSError as error:
        raise WorkbookDataError(f"无法读取工作簿「{source}」：{error}") from None


def _decimal(value: object, label: str) -> Decimal:
    if value in (None, ""):
        return Decimal(0)
    try:
        result = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise WorkbookDataError(f"字段「{label}」包含无法识别的金额：{value!r}。") from None
    if not result.is_finite():
        raise WorkbookDataError(f"字段「{label}」包含无法识别的金额：{value!r}。")
    return result


def _date(value: object, label: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.strip()).date()
        except ValueError:
            pass
    raise WorkbookDataError(f"字段「{label}」包含无法识别的日期：{value!r}。")


def _header_positions(
    headers: tuple[object, ...], required: set[str], sheet: str
) -> dict[str, int]:
    positions = {}
    for header in sorted(required):
        matches = [index for index, value in enumerate(headers) if value == header]
        if len(matches) > 1:
            raise WorkbookDataError(f"工作表「{sheet}」存在重复字段「{header}」。")
        if not matches:
            raise WorkbookDataError(f"工作表「{sheet}」缺少必填字段「{header}」。")
        positions[header] = matches[0]
    return positions


def _is_metadata(values: dict[str, object]) -> bool:
    return "公式" in values.values() and sum(value == "产品表" for value in values.values()) >= 4


def _validated_rows(
    cached_sheet: Worksheet,
    formula_sheet: Worksheet,
    required: set[str],
    numeric: set[str],
    date_fields: set[str],
    header_row: int,
    skip_row: Callable[[dict[str, object]], bool] | None = None,
) -> Iterator[dict[str, object]]:
    cached_rows = cached_sheet.iter_rows(values_only=True)
    formula_rows = formula_sheet.iter_rows()
    try:
        for _ in range(header_row):
            cached_headers = next(cached_rows)
            formula_headers = next(formula_rows)
    except StopIteration:
        if header_row == 1:
            raise WorkbookDataError(f"工作表「{cached_sheet.title}」为空，无法读取表头。") from None
        raise WorkbookDataError(
            f"工作表「{cached_sheet.title}」没有设置的第 {header_row} 行表头。"
        ) from None
    cached_positions = _header_positions(cached_headers, required, cached_sheet.title)
    formula_positions = _header_positions(
        tuple(cell.value for cell in formula_headers), numeric, formula_sheet.title
    )
    if any(cached_positions[field] != index for field, index in formula_positions.items()):
        raise WorkbookDataError(f"工作表「{cached_sheet.title}」的公式和值视图行结构不一致。")

    missing = object()
    for cached_row, formula_row in zip_longest(cached_rows, formula_rows, fillvalue=missing):
        if cached_row is missing or formula_row is missing:
            raise WorkbookDataError(f"工作表「{cached_sheet.title}」的公式和值视图行结构不一致。")
        values = {
            field: cached_row[index] if index < len(cached_row) else None
            for field, index in cached_positions.items()
        }
        if skip_row is not None and skip_row(values):
            continue
        active = any(values[field] not in (None, "", "未放款") for field in date_fields)
        if active:
            for field in numeric:
                cell = formula_row[formula_positions[field]]
                if cell.data_type == "f" and values[field] is None:
                    raise WorkbookDataError(
                        f"工作表「{cached_sheet.title}」字段「{field}」的公式没有缓存结果。"
                        "请用 Excel 或 WPS 重新计算后保存该工作簿，再重新导入。"
                    )
        yield values


def _open_workbook_views(path: Path):
    cached_book = _open_workbook(path, data_only=True)
    try:
        formula_book = _open_workbook(path, data_only=False)
    except (WorkbookDataError, ParseError, ValueError):
        cached_book.close()
        raise
    return cached_book, formula_book


def _sheet_pair(cached_book, formula_book, path: Path, sheet_name: str):
    try:
        return cached_book[sheet_name], formula_book[sheet_name]
    except KeyError:
        raise WorkbookDataError(f"工作簿「{path}」缺少工作表「{sheet_name}」。") from None


def _text(value: object) -> str | None:
    if value is None:
        return None
    result = str(value)
    return result if result.strip() else None


def _read_operations(path: Path, settings: SourceSettings) -> list[OperationalRecord]:
    cached_book, formula_book = _open_workbook_views(path)
    try:
        cached, formula = _sheet_pair(cached_book, formula_book, path, settings.operations_sheet)
        records = []
        for values in _validated_rows(
            cached,
            formula,
            settings.operation_fields(),
            settings.operation_numeric_fields(),
            settings.operation_date_fields(),
            settings.operations_header_row,
            _is_metadata,
        ):
            present_dates = [
                _date(values[field], field)
                for field in settings.operation_date_fields()
                if values[field] not in (None, "")
            ]
            if not present_dates:
                continue
            for field in settings.operation_numeric_fields():
                _decimal(values[field], field)
            records.append(
                OperationalRecord(
                    _text(values.get(settings.project_count.value_field)),
                    _text(values.get(settings.scatter_count.filters[0].field)),
                    _text(values.get("目的口岸")),
                    present_dates[0],
                    _text(values.get("B1供应商")),
                    _decimal(
                        values.get(settings.business_total.sales.value_field),
                        settings.business_total.sales.value_field,
                    ),
                    _decimal(
                        values.get(settings.project_profit.value_field),
                        settings.project_profit.value_field,
                    ),
                    values,
                )
            )
        return records
    finally:
        formula_book.close()
        cached_book.close()


def _read_funds(
    path: Path, allowed_years: Iterable[int], settings: SourceSettings
) -> list[FundRecord]:
    years = sorted(set(allowed_years))
    if "{年份}" in settings.funds_sheet and not years:
        raise WorkbookDataError("未指定资金工作表年份。")
    cached_book, formula_book = _open_workbook_views(path)
    try:
        candidates = (
            [settings.funds_sheet.replace("{年份}", str(year)) for year in years]
            if "{年份}" in settings.funds_sheet
            else [settings.funds_sheet]
        )
        selected = [name for name in candidates if name in cached_book.sheetnames]
        if not selected:
            names = "、".join(f"「{name}」" for name in candidates)
            raise WorkbookDataError(f"工作簿「{path}」未找到设置的资金工作表：{names}。")
        records = []
        for sheet_name in selected:
            cached, formula = _sheet_pair(cached_book, formula_book, path, sheet_name)
            for values in _validated_rows(
                cached,
                formula,
                settings.fund_fields(),
                settings.fund_numeric_fields(),
                settings.fund_date_fields(),
                settings.funds_header_row,
            ):
                present_dates = [
                    _date(values[field], field)
                    for field in settings.fund_date_fields()
                    if values[field] not in (None, "", "未放款")
                ]
                if not present_dates:
                    continue
                for field in settings.fund_numeric_fields():
                    _decimal(values[field], field)
                records.append(
                    FundRecord(
                        _text(values.get(settings.fund_profit.channel_field)) or "",
                        present_dates[0],
                        _decimal(
                            values.get(settings.fund_profit.amount_field),
                            settings.fund_profit.amount_field,
                        ),
                        _decimal(
                            values.get(settings.fund_profit.operation_fee_field),
                            settings.fund_profit.operation_fee_field,
                        ),
                        values,
                    )
                )
        return records
    finally:
        formula_book.close()
        cached_book.close()


def read_operations(
    path: Path, settings: SourceSettings = DEFAULT_SOURCE_SETTINGS
) -> list[OperationalRecord]:
    settings.validate()
    source = Path(path)
    try:
        return _read_operations(source, settings)
    except WorkbookDataError:
        raise
    except (BadZipFile, KeyError, ParseError, ValueError):
        raise WorkbookDataError(f"文件「{source}」不是有效的 XLSX 工作簿。") from None


def read_funds(
    path: Path,
    allowed_years: Iterable[int],
    settings: SourceSettings = DEFAULT_SOURCE_SETTINGS,
) -> list[FundRecord]:
    settings.validate()
    source = Path(path)
    try:
        return _read_funds(source, allowed_years, settings)
    except WorkbookDataError:
        raise
    except (BadZipFile, KeyError, ParseError, ValueError):
        raise WorkbookDataError(f"文件「{source}」不是有效的 XLSX 工作簿。") from None
