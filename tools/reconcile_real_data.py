import hashlib
import json
import sys
import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ledger_reporter.services.history import HistoryRepository
from ledger_reporter.services.report_service import ReportService

START = date(2026, 8, 1)
END = date(2026, 8, 6)
ZERO = Decimal(0)
CAPITAL_COST = Decimal("0.0448")
FUND_RATES = {
    "广州美鑫通国际供应链有限公司": Decimal("0.10"),
    "浙江飞速供应链管理有限公司": Decimal("0.12"),
}


def _fingerprint(path: Path) -> tuple[int, int, str]:
    stat = path.stat()
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return stat.st_size, stat.st_mtime_ns, digest.hexdigest()


def _as_date(value: object, epoch) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        return converted.date() if isinstance(converted, datetime) else converted
    if isinstance(value, str):
        return datetime.fromisoformat(value.strip()).date()
    raise ValueError(f"无法识别日期：{value!r}")


def _decimal(value: object) -> Decimal:
    return ZERO if value in (None, "") else Decimal(str(value))


def _header_map(row: tuple[object, ...]) -> dict[str, int]:
    return {str(value): index for index, value in enumerate(row) if value is not None}


def _direct_operations(path: Path) -> dict[str, int | Decimal]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet = workbook["台账明细"]
        rows = sheet.iter_rows(values_only=True)
        columns = _header_map(next(rows))
        project_count = 0
        project_profit = ZERO
        scatter_count = 0
        scatter_profit = ZERO
        for row in rows:
            raw_date = row[columns["预计起飞时间"]]
            if raw_date in (None, ""):
                continue
            if row[columns["预估总应收"]] == "公式" and all(
                row[columns[header]] == "产品表"
                for header in ("提单号", "目的口岸", "预计起飞时间", "B1供应商")
            ):
                continue
            departure = _as_date(raw_date, workbook.epoch)
            if not START <= departure <= END:
                continue
            bill_count = int(row[columns["提单号"]] not in (None, ""))
            profit = _decimal(row[columns["预估毛利润"]])
            if row[columns["项目类型"]] == "散采":
                scatter_count += bill_count
                scatter_profit += profit
            else:
                project_count += bill_count
                project_profit += profit
        return {
            "project_count": project_count,
            "project_profit": project_profit,
            "scatter_count": scatter_count,
            "scatter_profit": scatter_profit,
        }
    finally:
        workbook.close()


def _direct_funds(path: Path) -> dict[str, Decimal]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        amount = ZERO
        profit = ZERO
        for sheet_name in (
            name for name in workbook.sheetnames if name.startswith("资金散板汇总2026")
        ):
            rows = workbook[sheet_name].iter_rows(values_only=True)
            columns = _header_map(next(rows))
            for row in rows:
                raw_date = row[columns["信容付款日期"]]
                if raw_date in (None, "", "未放款"):
                    continue
                payment_date = _as_date(raw_date, workbook.epoch)
                if not START <= payment_date <= END:
                    continue
                channel = str(row[columns["渠道名称"]] or "")
                row_amount = _decimal(row[columns["付款金额合计（90%）"]])
                amount += row_amount
                profit += row_amount * (FUND_RATES[channel] - CAPITAL_COST) * Decimal(60) / Decimal(
                    365
                ) + _decimal(row[columns["应收操作费"]])
        return {"fund_amount": amount, "fund_profit": profit}
    finally:
        workbook.close()


def reconcile(
    funds_path: Path,
    operations_path: Path,
    history_path: Path,
) -> dict[str, object]:
    expected = _direct_operations(operations_path) | _direct_funds(funds_path)
    bundle = ReportService(HistoryRepository(history_path)).generate(
        funds_path,
        operations_path,
        date(2026, 8, 10),
    )
    snapshot = next(
        item for item in bundle.weeks if item.period.start == START and item.period.end == END
    )
    actual = {
        "project_count": snapshot.metrics.project_count,
        "project_profit": snapshot.metrics.project_profit,
        "scatter_count": snapshot.metrics.scatter_count,
        "scatter_profit": snapshot.metrics.scatter_profit,
        "fund_amount": snapshot.metrics.fund_amount,
        "fund_profit": snapshot.metrics.fund_profit,
    }
    differences = {name: actual[name] - expected[name] for name in expected}
    return {"expected": expected, "actual": actual, "differences": differences}


def _json_value(value: object) -> object:
    return str(value) if isinstance(value, Decimal) else value


def main(argv: list[str] | None = None) -> int:
    arguments = sys.argv[1:] if argv is None else argv
    if len(arguments) != 2:
        print("用法: python tools/reconcile_real_data.py <资金台账.xlsx> <运营台账.xlsx>")
        return 2
    funds_path, operations_path = map(Path, arguments)
    sources = (funds_path, operations_path)
    before = {path: _fingerprint(path) for path in sources}
    with tempfile.TemporaryDirectory() as directory:
        result = reconcile(funds_path, operations_path, Path(directory) / "history.sqlite3")
    after = {path: _fingerprint(path) for path in sources}
    if after != before:
        print("对账期间源工作簿发生变化。", file=sys.stderr)
        return 1
    output = {
        "period": [START.isoformat(), END.isoformat()],
        **{
            section: {name: _json_value(value) for name, value in values.items()}
            for section, values in result.items()
        },
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))
    return int(any(value != 0 for value in result["differences"].values()))


if __name__ == "__main__":
    raise SystemExit(main())
