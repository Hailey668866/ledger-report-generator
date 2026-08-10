import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def _label(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return f"{value.year}年{value.month}月"
    return str(value or "")


def extract_baseline(source: Path, target: Path) -> None:
    if source.resolve() == target.resolve() or (target.exists() and source.samefile(target)):
        raise ValueError("source and target must refer to different files")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook["Sheet1"]
        rows = [
            {"label": _label(row[0]), "values": list(row[1:])}
            for row in worksheet.iter_rows(
                min_row=3,
                max_row=28,
                min_col=1,
                max_col=11,
                values_only=True,
            )
        ]
    finally:
        workbook.close()

    payload = {
        "version": "fy2026-requirement-2026-08-10-v1",
        "fiscal_year": 2026,
        "frozen_through": "2026-07-31",
        "rows": rows,
    }
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("target", type=Path)
    args = parser.parse_args()
    extract_baseline(args.source, args.target)


if __name__ == "__main__":
    main()
