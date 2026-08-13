from contextlib import closing
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageStat

from ledger_reporter.domain.models import FundRecord, OperationalRecord
from ledger_reporter.exporters.excel import export_excel
from ledger_reporter.exporters.png import export_pngs
from ledger_reporter.services.history import HistoryRepository
from ledger_reporter.services.report_service import ReportService


def test_generates_persists_and_exports_both_tables(tmp_path: Path) -> None:
    history_path = tmp_path / "history.sqlite3"
    repository = HistoryRepository(history_path)
    service = ReportService(repository)
    operations = [
        OperationalRecord(
            "001",
            "BSA",
            "LAX",
            date(2026, 8, 3),
            "其他供应商",
            Decimal(1000),
            Decimal(100),
        ),
        OperationalRecord(
            "002",
            "散采",
            "OSL",
            date(2026, 8, 4),
            "其他供应商",
            Decimal(500),
            Decimal(20),
        ),
    ]
    funds = [
        FundRecord(
            "广州美鑫通国际供应链有限公司",
            date(2026, 8, 4),
            Decimal(900),
            Decimal(5),
        )
    ]

    bundle = service.generate_from_records(date(2026, 8, 7), operations, funds)
    excel_path = tmp_path / "2026财年台账报表.xlsx"
    export_excel(bundle, excel_path)
    png_paths = export_pngs(bundle, tmp_path / "png")

    reloaded = HistoryRepository(history_path).load_weeks(2026)
    assert [(week.period.start, week.period.end) for week in reloaded] == [
        (date(2026, 8, 1), date(2026, 8, 6))
    ]
    assert reloaded[0].metrics.project_count == 1
    assert reloaded[0].metrics.scatter_count == 1
    assert reloaded[0].metrics.fund_amount == Decimal(900)

    with closing(load_workbook(excel_path, data_only=False)) as workbook:
        assert workbook.sheetnames == ["经营汇总", "自营项目周报"]
        assert workbook["经营汇总"]["A1"].value == "日期"
        assert "自营项目数据情况" in workbook["自营项目周报"]["A1"].value

    assert [path.name for path in png_paths] == ["经营汇总.png", "自营项目周报.png"]
    for path in png_paths:
        with Image.open(path) as image:
            assert image.format == "PNG"
            assert image.width >= 1000
            assert image.height >= 400
            assert sum(ImageStat.Stat(image.convert("RGB")).var) > 0
