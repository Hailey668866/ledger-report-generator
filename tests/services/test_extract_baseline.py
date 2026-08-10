import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from tools.extract_baseline import extract_baseline


def _create_source(path: Path) -> None:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        for row in range(1, 29):
            for column in range(1, 12):
                worksheet.cell(row, column, f"{row}-{column}")
        workbook.save(path)
    finally:
        workbook.close()


def test_rejects_source_as_target_without_changing_workbook(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _create_source(source)
    original = source.read_bytes()

    with pytest.raises(ValueError, match=r"(?i)source.*target|target.*source"):
        extract_baseline(source, source)

    assert source.read_bytes() == original
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["Sheet1"]
    finally:
        workbook.close()


def test_rejects_existing_target_that_aliases_source(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    target = tmp_path / "alias.xlsx"
    _create_source(source)
    target.hardlink_to(source)
    original = source.read_bytes()

    with pytest.raises(ValueError, match=r"(?i)source.*target|target.*source"):
        extract_baseline(source, target)

    assert source.read_bytes() == original


def test_extracts_rows_three_through_twenty_eight_to_a_different_target(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    target = tmp_path / "output" / "baseline.json"
    _create_source(source)
    original = source.read_bytes()

    extract_baseline(source, target)

    payload = json.loads(target.read_text(encoding="utf-8"))
    assert source.read_bytes() == original
    assert len(payload["rows"]) == 26
    assert payload["rows"][0] == {
        "label": "3-1",
        "values": [f"3-{column}" for column in range(2, 12)],
    }
    assert payload["rows"][-1]["label"] == "28-1"
