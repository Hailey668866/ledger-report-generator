from dataclasses import replace
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula

from ledger_reporter.io import workbooks
from ledger_reporter.io.errors import WorkbookDataError
from ledger_reporter.io.source_settings import DEFAULT_SOURCE_SETTINGS, _defaults
from ledger_reporter.io.workbooks import read_funds, read_operations

OPS_HEADERS = (
    "提单号",
    "项目类型",
    "目的口岸",
    "预计起飞时间",
    "B1供应商",
    "预估总应收",
    "预估毛利润",
)
FUND_HEADERS = ("渠道名称", "信容付款日期", "付款金额合计（90%）", "应收操作费")


def save_book(
    path: Path,
    sheet_name: str,
    headers: tuple[str, ...],
    rows: list[tuple[object, ...]],
    *,
    header_row: int = 1,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for _ in range(header_row - 1):
        worksheet.append(("导出说明",))
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def corrupt_first_sheet_xml(path: Path) -> None:
    with ZipFile(path) as archive:
        members = {name: archive.read(name) for name in archive.namelist()}
    members["xl/worksheets/sheet1.xml"] += b"<"
    with ZipFile(path, "w") as archive:
        for name, contents in members.items():
            archive.writestr(name, contents)


def test_read_operations_maps_headers_and_converts_values(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    headers = (
        "预计起飞时间",
        "预估毛利润",
        "目的口岸",
        "B1供应商",
        "提单号",
        "预估总应收",
        "项目类型",
    )
    save_book(
        path,
        "台账明细",
        headers,
        [
            (
                datetime(2026, 8, 2, 10, 30),  # noqa: DTZ001
                12.5,
                "LAX",
                "供应商 A",
                "000123",
                100,
                "散板",
            )
        ],
    )

    records = read_operations(path)

    assert len(records) == 1
    record = records[0]
    assert record.bill_no == "000123"
    assert record.project_type == "散板"
    assert record.destination == "LAX"
    assert record.departure == date(2026, 8, 2)
    assert record.supplier == "供应商 A"
    assert record.receivable == Decimal(100)
    assert record.gross_profit == Decimal("12.5")


def test_read_operations_uses_custom_sheet_second_header_row_and_renamed_headers(
    tmp_path: Path,
) -> None:
    path = tmp_path / "operations.xlsx"
    settings = _defaults(
        {
            "operations_sheet": "运营数据",
            "operations_header_row": 2,
            "operations_bill_no": "运单",
            "operations_project_type": "类别",
            "operations_destination": "到达港",
            "operations_departure": "起飞日",
            "operations_supplier": "服务商",
            "operations_receivable": "应收款",
            "operations_gross_profit": "毛利",
        }
    )
    save_book(
        path,
        "运营数据",
        ("毛利", "到达港", "运单", "应收款", "服务商", "起飞日", "类别"),
        [(12.5, "LAX", "000123", 100, "供应商 A", "2026-08-02", "散板")],
        header_row=2,
    )

    records = read_operations(path, settings)

    assert len(records) == 1
    assert records[0].values["运单"] == "000123"
    assert records[0].values["到达港"] == "LAX"
    assert records[0].values["毛利"] == 12.5


def test_read_operations_accepts_iso_datetime_string(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("B-1", "散板", "LAX", "2026-08-02 10:30:00", "供应商 A", 100, 10)],
    )

    records = read_operations(path)

    assert records[0].departure == date(2026, 8, 2)


def test_read_operations_skips_embedded_product_table_metadata_row(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [
            ("产品表", "运费计费重时更新", "产品表", "产品表", "产品表", "公式", None),
            ("B-1", "散板", "LAX", "2026-08-02", "供应商 A", 100, 10),
        ],
    )

    records = read_operations(path)

    assert [record.bill_no for record in records] == ["B-1"]


def test_product_table_metadata_skips_uncached_amount_formula(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("产品表", "运费计费重时更新", "产品表", "产品表", "产品表", "公式", "=1+1")],
    )

    assert read_operations(path) == []


def test_read_operations_rejects_bad_date_in_regular_business_row(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("B-1", "散板", "LAX", "待确认", "供应商 A", 100, 10)],
    )

    with pytest.raises(WorkbookDataError, match="预计起飞时间.*待确认"):
        read_operations(path)


def test_read_operations_bad_date_names_custom_header(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    settings = _defaults({"operations_departure": "计划起飞日期"})
    headers = tuple(
        "计划起飞日期" if header == "预计起飞时间" else header for header in OPS_HEADERS
    )
    save_book(
        path,
        "台账明细",
        headers,
        [("B-1", "散板", "LAX", "待确认", "供应商 A", 100, 10)],
    )

    with pytest.raises(WorkbookDataError, match="计划起飞日期.*待确认"):
        read_operations(path, settings)


def test_read_operations_does_not_skip_product_markers_without_formula_metadata(
    tmp_path: Path,
) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("产品表", "项目", "产品表", "产品表", "产品表", 100, 10)],
    )

    with pytest.raises(WorkbookDataError, match="预计起飞时间.*产品表"):
        read_operations(path)


def test_read_operations_rejects_missing_required_header(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(path, "台账明细", OPS_HEADERS[:-1], [])

    with pytest.raises(WorkbookDataError, match="预估毛利润"):
        read_operations(path)


def test_read_operations_missing_custom_header_names_the_configured_header(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(path, "台账明细", OPS_HEADERS, [])
    settings = _defaults({"operations_bill_no": "自定义提单"})

    with pytest.raises(WorkbookDataError, match="缺少必填字段.*自定义提单"):
        read_operations(path, settings)


def test_read_operations_rejects_header_row_beyond_sheet_data(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(path, "台账明细", ("导出说明",), [])

    with pytest.raises(WorkbookDataError, match="没有设置的第 2 行表头"):
        read_operations(path, replace(DEFAULT_SOURCE_SETTINGS, operations_header_row=2))


def test_read_operations_rejects_duplicate_required_header(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(path, "台账明细", OPS_HEADERS + ("提单号",), [])

    with pytest.raises(WorkbookDataError, match="重复字段.*提单号"):
        read_operations(path)


def test_read_operations_rejects_corrupt_xlsx(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    path.write_bytes(b"not an xlsx workbook")

    with pytest.raises(WorkbookDataError, match="不是有效的 XLSX"):
        read_operations(path)


@pytest.mark.parametrize("value", ["NaN", "Infinity"])
def test_read_operations_rejects_non_finite_amounts(tmp_path: Path, value: str) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("B-1", "散板", "LAX", "2026-08-02", "供应商 A", value, 10)],
    )

    with pytest.raises(WorkbookDataError, match="预估总应收"):
        read_operations(path)


def test_read_operations_rejects_zip_without_ooxml_parts(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    with ZipFile(path, "w") as archive:
        archive.writestr("dummy.txt", "not an OOXML workbook")

    with pytest.raises(WorkbookDataError, match="不是有效的 XLSX"):
        read_operations(path)


def test_read_operations_rejects_lazy_worksheet_xml_corruption(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("B-1", "散板", "LAX", "2026-08-02", "供应商 A", 100, 10)],
    )
    corrupt_first_sheet_xml(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    workbook.close()

    with pytest.raises(WorkbookDataError, match="不是有效的 XLSX"):
        read_operations(path)


def test_read_funds_reads_requested_year_sheet(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(
        path,
        "资金散板汇总2026",
        FUND_HEADERS,
        [("渠道 A", "2026-08-05", 90, 3.25)],
    )

    records = read_funds(path, {2026})

    assert len(records) == 1
    record = records[0]
    assert record.channel == "渠道 A"
    assert record.payment_date == date(2026, 8, 5)
    assert record.amount == Decimal(90)
    assert record.operation_fee == Decimal("3.25")


def test_read_funds_uses_custom_year_template_and_renamed_headers(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    settings = _defaults(
        {
            "funds_sheet": "付款明细-{年份}",
            "funds_channel": "付款渠道",
            "funds_payment_date": "付款日",
            "funds_amount": "付款额",
            "funds_operation_fee": "操作费",
        }
    )
    save_book(
        path,
        "付款明细-2026",
        ("操作费", "付款额", "付款渠道", "付款日"),
        [(3.25, 90, "渠道 A", "2026-08-05")],
    )

    records = read_funds(path, {2026, 2027}, settings)

    assert [(record.channel, record.amount) for record in records] == [("渠道 A", Decimal(90))]


def test_read_funds_reads_static_configured_sheet_once_for_multiple_years(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(path, "跨年资金", FUND_HEADERS, [("渠道 A", "2026-08-05", 90, 3.25)])

    records = read_funds(
        path, {2026, 2027}, replace(DEFAULT_SOURCE_SETTINGS, funds_sheet="跨年资金")
    )

    assert [record.channel for record in records] == ["渠道 A"]


def test_read_funds_static_configured_sheet_allows_empty_years(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(path, "跨年资金", FUND_HEADERS, [("渠道 A", "2026-08-05", 90, 3.25)])

    records = read_funds(path, set(), replace(DEFAULT_SOURCE_SETTINGS, funds_sheet="跨年资金"))

    assert [record.channel for record in records] == ["渠道 A"]


def test_read_funds_year_template_rejects_empty_allowed_years(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(path, "资金散板汇总2026", FUND_HEADERS, [])

    with pytest.raises(WorkbookDataError, match="未指定资金工作表年份"):
        read_funds(path, set())


def test_read_funds_missing_configured_candidates_lists_exact_sheet_names(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(path, "其他工作表", FUND_HEADERS, [])
    settings = replace(DEFAULT_SOURCE_SETTINGS, funds_sheet="付款明细-{年份}")

    with pytest.raises(WorkbookDataError) as error:
        read_funds(path, {2027, 2026}, settings)

    assert "付款明细-2026" in str(error.value)
    assert "付款明细-2027" in str(error.value)


def test_read_funds_uses_allowed_years_parameter(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(
        path,
        "资金散板汇总2026",
        FUND_HEADERS,
        [("渠道 A", "2026-08-05", 90, 3.25)],
    )

    records = read_funds(path, allowed_years={2026, 2027})

    assert len(records) == 1


def test_read_funds_reuses_two_workbook_views_across_selected_sheets(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(
        path,
        "资金散板汇总2026",
        FUND_HEADERS,
        [("渠道 A", "2026-08-05", 90, 3.25)],
    )
    workbook = load_workbook(path)
    worksheet = workbook.create_sheet("资金散板汇总2027")
    worksheet.append(FUND_HEADERS)
    worksheet.append(("渠道 B", "2027-04-05", 80, 2.25))
    workbook.save(path)
    workbook.close()

    calls: list[bool] = []
    open_workbook = workbooks._open_workbook

    def spy_open_workbook(source: Path, data_only: bool):
        calls.append(data_only)
        return open_workbook(source, data_only)

    monkeypatch.setattr(workbooks, "_open_workbook", spy_open_workbook)

    records = workbooks.read_funds(path, allowed_years={2026, 2027})

    assert {record.channel for record in records} == {"渠道 A", "渠道 B"}
    assert calls == [True, False]


def test_read_operations_skips_formula_cache_check_for_blank_departure(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("B-1", "散板", "LAX", None, "供应商 A", 100, "=1+1")],
    )

    assert read_operations(path) == []


def test_read_funds_skips_formula_cache_check_for_blank_payment_date(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(
        path,
        "资金散板汇总2026",
        FUND_HEADERS,
        [("渠道 A", None, "=1+1", 3)],
    )

    assert read_funds(path, {2026}) == []


def test_read_funds_skips_rows_that_are_explicitly_not_disbursed(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(
        path,
        "资金散板汇总2026",
        FUND_HEADERS,
        [
            ("渠道 A", "未放款", 90, 3.25),
            ("渠道 A", "2026-08-05", 80, 2.25),
        ],
    )

    records = read_funds(path, {2026})

    assert [record.amount for record in records] == [Decimal(80)]


def test_not_disbursed_row_skips_uncached_amount_formulas(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(
        path,
        "资金散板汇总2026",
        FUND_HEADERS,
        [("渠道 A", "未放款", "=1+1", "=2+2")],
    )

    assert read_funds(path, {2026}) == []


def test_read_funds_rejects_other_bad_payment_dates(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    save_book(
        path,
        "资金散板汇总2026",
        FUND_HEADERS,
        [("渠道 A", "日期待确认", 90, 3.25)],
    )

    with pytest.raises(WorkbookDataError, match="信容付款日期.*日期待确认"):
        read_funds(path, {2026})


def test_read_funds_bad_amount_names_custom_header(tmp_path: Path) -> None:
    path = tmp_path / "funds.xlsx"
    settings = _defaults({"funds_amount": "实付金额"})
    headers = tuple(
        "实付金额" if header == "付款金额合计（90%）" else header for header in FUND_HEADERS
    )
    save_book(
        path,
        "资金散板汇总2026",
        headers,
        [("渠道 A", "2026-08-05", "金额待确认", 3.25)],
    )

    with pytest.raises(WorkbookDataError, match="实付金额.*金额待确认"):
        read_funds(path, {2026}, settings)


def test_read_operations_rejects_formula_without_cached_result(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("B-1", "散板", "LAX", "2026-08-02", "供应商 A", 100, "=1+1")],
    )

    with pytest.raises(WorkbookDataError, match="公式没有缓存结果"):
        read_operations(path)


def test_read_operations_checks_formula_cache_in_custom_renamed_column(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    settings = _defaults(
        {
            "operations_receivable": "自定义应收",
            "operations_gross_profit": "自定义毛利",
        }
    )
    save_book(
        path,
        "台账明细",
        ("自定义毛利", "预计起飞时间", "提单号", "项目类型", "目的口岸", "B1供应商", "自定义应收"),
        [("=1+1", "2026-08-02", "B-1", "散板", "LAX", "供应商 A", 100)],
    )

    with pytest.raises(WorkbookDataError, match="自定义毛利.*公式没有缓存结果"):
        read_operations(path, settings)


def test_public_readers_validate_settings_before_opening_workbook(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"

    with pytest.raises(ValueError, match="正整数"):
        read_operations(path, replace(DEFAULT_SOURCE_SETTINGS, operations_header_row=0))
    with pytest.raises(ValueError, match="正整数"):
        read_funds(path, {2026}, replace(DEFAULT_SOURCE_SETTINGS, funds_header_row=0))


def test_read_operations_rejects_array_formula_without_cached_result(tmp_path: Path) -> None:
    path = tmp_path / "operations.xlsx"
    save_book(
        path,
        "台账明细",
        OPS_HEADERS,
        [("B-1", "散板", "LAX", "2026-08-02", "供应商 A", 100, 10)],
    )
    workbook = load_workbook(path)
    workbook["台账明细"]["G2"] = ArrayFormula(ref="G2", text="=1+1")
    workbook.save(path)
    workbook.close()

    with pytest.raises(WorkbookDataError, match="公式没有缓存结果"):
        read_operations(path)
