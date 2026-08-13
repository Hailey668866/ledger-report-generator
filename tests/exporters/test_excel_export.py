from contextlib import closing
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ledger_reporter.exporters.excel import export_excel
from ledger_reporter.services.baseline import load_fy2026_baseline


def _row_for_label(sheet: Worksheet, label: str) -> int:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == label:
            return row
    raise AssertionError(f"missing row label: {label}")


def test_exports_only_two_formal_sheets_without_source_notes_or_errors(
    report_bundle,
    tmp_path: Path,
) -> None:
    output = tmp_path / "nested" / "2026财年台账报表.xlsx"

    export_excel(report_bundle, output)

    assert output.is_file()
    with closing(load_workbook(output, data_only=False)) as book:
        assert book.sheetnames == ["经营汇总", "自营项目周报"]
        assert book["经营汇总"]["A1"].value == "日期"
        cells = [cell for sheet in book.worksheets for row in sheet.iter_rows() for cell in row]
        assert "数据源" not in {cell.value for cell in cells}
        assert not any(cell.data_type == "e" for cell in cells)
        assert not any(isinstance(cell.value, str) and cell.value.startswith("#") for cell in cells)


def test_summary_sheet_aggregates_dynamic_week_into_month_and_quarter(
    report_bundle,
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"
    export_excel(report_bundle, output)

    with closing(load_workbook(output, data_only=False)) as book:
        sheet = book["经营汇总"]
        quarter_row = _row_for_label(sheet, "Q2(26.7-26.9)")
        month_row = _row_for_label(sheet, "2026年8月")
        week_row = _row_for_label(sheet, "W1（8.1-8.6）")

        expected_metrics = [2, 100, 1, 20, 900, 5, 0, 0, 125]
        assert [sheet.cell(quarter_row, column).value for column in range(2, 11)] == (
            expected_metrics
        )
        assert [sheet.cell(month_row, column).value for column in range(2, 11)] == (
            expected_metrics
        )
        assert [sheet.cell(week_row, column).value for column in range(2, 11)] == (expected_metrics)
        assert sheet.cell(quarter_row, 11).value == 5_000_000
        assert sheet.cell(month_row, 11).value == 1_670_000
        assert sheet.cell(week_row, 11).value is None
        assert sheet.cell(quarter_row, 12).value == pytest.approx(125 / 5_000_000)
        assert sheet.cell(month_row, 12).value == pytest.approx(125 / 1_670_000)
        assert sheet.cell(week_row, 12).value is None


def test_dynamic_month_updates_existing_quarter_without_changing_frozen_month(
    report_bundle,
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"
    bundle = replace(
        report_bundle,
        baseline_rows=load_fy2026_baseline().rows,
    )
    export_excel(bundle, output)

    with closing(load_workbook(output, data_only=False)) as book:
        sheet = book["经营汇总"]
        labels = [sheet.cell(row, 1).value for row in range(1, sheet.max_row + 1)]
        assert labels.count("Q2(26.7-26.9)") == 1
        quarter_row = _row_for_label(sheet, "Q2(26.7-26.9)")
        july_row = _row_for_label(sheet, "2026年7月")
        august_row = _row_for_label(sheet, "2026年8月")

        assert sheet.cell(quarter_row, 2).value == 837
        assert sheet.cell(quarter_row, 10).value == pytest.approx(-2_118_033.69446279 + 125)
        assert sheet.cell(july_row, 2).value == 835
        assert sheet.cell(july_row, 10).value == pytest.approx(-2_118_033.69446279)
        assert august_row > july_row


def test_frozen_baseline_total_is_preserved_instead_of_recomputed(
    report_bundle,
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"
    frozen_row = {
        "label": "2026年7月",
        "values": [1, 10, 0, 0, 0, 0, 0, 0, 999, 1_670_000],
    }
    bundle = replace(report_bundle, baseline_rows=(frozen_row,), weeks=())
    export_excel(bundle, output)

    with closing(load_workbook(output, data_only=False)) as book:
        sheet = book["经营汇总"]
        july_row = _row_for_label(sheet, "2026年7月")
        assert sheet.cell(july_row, 10).value == 999
        assert sheet.cell(july_row, 12).value == pytest.approx(999 / 1_670_000)


def test_business_sheet_preserves_numeric_rates_and_uses_ten_thousand_units(
    report_bundle,
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"
    export_excel(report_bundle, output)

    with closing(load_workbook(output, data_only=False)) as book:
        sheet = book["自营项目周报"]
        assert sheet["A1"].value == ("2026年8月W1（8.1-8.6）自营项目数据情况，单位：万元")
        assert sheet["C3"].value == pytest.approx(0.0102)
        assert sheet["C3"].number_format == "0.00%"
        assert sheet["C6"].value == "固定差价2%"
        assert sheet["E3"].value == pytest.approx(0.01)
        assert sheet["F3"].value == pytest.approx(0.1)
        assert sheet["A14"].value == "销售额合计：1"
        assert sheet["D14"].value == 12
        assert sheet["E14"].value == pytest.approx(0.12)
        assert sheet["F14"].value == pytest.approx(0.1)
        assert sheet["D14"].number_format == "#,##0"
        assert sheet["E14"].number_format == "#,##0.00"
        assert sheet["F14"].number_format == "0.00%"


def test_export_applies_report_layout_and_template_palette(
    report_bundle,
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"
    export_excel(report_bundle, output)

    with closing(load_workbook(output, data_only=False)) as book:
        summary = book["经营汇总"]
        business = book["自营项目周报"]

        assert summary.sheet_view.showGridLines is False
        assert business.sheet_view.showGridLines is False
        assert summary.freeze_panes == "B3"
        assert business.freeze_panes == "A3"
        assert {str(item) for item in summary.merged_cells.ranges} == {
            "A1:A2",
            "B1:C1",
            "D1:E1",
            "F1:G1",
            "H1:I1",
            "J1:J2",
            "K1:K2",
            "L1:L2",
        }
        assert {str(item) for item in business.merged_cells.ranges} == {"A1:F1"}
        assert summary["B1"].fill.fgColor.rgb == "0091AADD"
        assert business["A1"].fill.fgColor.rgb == "0060708E"
        assert business["A1"].font.color.rgb == "00FFFFFF"
        assert summary["C3"].number_format == "#,##0.00"
        assert summary["L4"].number_format == "0%"
